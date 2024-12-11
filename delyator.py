import concurrent.futures
import logging
import os
import time
from concurrent.futures import ThreadPoolExecutor
from tkinter import messagebox

from openpyxl import load_workbook, Workbook

from ExcelOperations import ExcelOperations
from FileManager import FileManager as fm
from TKinter import TKinter

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def sort_group(file_path, sheet, folder_column, file_column):
    logging.info(f"Начало сортировки данных по столбцу {file_column}")
    main_workbook = sheet.parent
    new_sheet_title = f'sorted_{sheet.title}'
    if new_sheet_title in main_workbook.sheetnames:
        logging.info(f"Лист '{new_sheet_title}' существует. Удаляем его.")
        del main_workbook[new_sheet_title]

    new_sheet = main_workbook.create_sheet(new_sheet_title)

    rows_styles = []
    for row in sheet.iter_rows():
        row_data = [(cell.value, cell) for cell in row]
        rows_styles.append(row_data)

    header = rows_styles[0]
    data = rows_styles[1:]
    sorted_data = sorted(data, key=lambda x: (
        x[file_column][0] if x[file_column][0] is not None else "",
        x[folder_column][0] if x[folder_column][0] is not None else ""
    ))

    for i, row_data in enumerate([header] + sorted_data, start=1):
        for j, (value, cell) in enumerate(row_data, start=1):
            new_cell = new_sheet.cell(row=i, column=j)
            new_cell.value = value
            if cell.has_style:
                ExcelOperations.apply_cell_styles(cell, new_cell)
    main_workbook.save(file_path)
    logging.info(f"Сортировка завершена. Данные сохранены в новый лист '{new_sheet_title}'")

    return new_sheet

def process_group(sheet, folder_name, files, output_folder):
    logging.info(f"Начинается обработка папки {folder_name}")

    sanitized_folder_name = fm.sanitize_filename(folder_name)
    folder_path = os.path.join(output_folder, sanitized_folder_name)

    os.makedirs(folder_path, exist_ok=True)
    logging.info(f"Папка создана или уже существует: {folder_path}")

    for file_name, rows in files.items():
        sanitized_file_name = fm.sanitize_filename(file_name)
        file_path = os.path.join(folder_path, f'{sanitized_file_name}.xlsx')

        new_wb = Workbook()
        new_sheet = new_wb.active
        dims = ExcelOperations.size_dims(sheet)
        ExcelOperations.create_header(sheet, new_sheet)
        ExcelOperations.set_columns_width(new_sheet, dims)

        row_index = 2

        for row in rows:
            ExcelOperations.copy_row(sheet, new_sheet, row[0].row, row_index)
            row_index += 1


        new_wb.save(file_path)
        logging.info(f"Данные для начальника {file_name} сохранены в файл: {file_path}")

def process_multi(sheet, gen_column, spc_column, output_folder):
    logging.info("Группировка данных по начальнику и куратору")
    grouped_data = {}
    total_row = len(list(sheet.iter_rows(min_row=2)))
    processed_rows = 0

    for row in sheet.iter_rows(min_row=2):
        folder_name = row[gen_column].value
        file_name = row[spc_column].value
        if not folder_name or not file_name:
            continue

        if folder_name not in grouped_data:
            grouped_data[folder_name] = {}

        if file_name not in grouped_data[folder_name]:
            grouped_data[folder_name][file_name] = []

        grouped_data[folder_name][file_name].append(row)
        # processed_rows += 1
        #
        # if processed_rows % (total_row // 10) == 0:
        #     progress = int((processed_rows / total_row) * 100)
        #     update_progress(progress)

    tasks = [(sheet, folder_name, files, output_folder) for folder_name, files in grouped_data.items()]
    with ThreadPoolExecutor() as executor:
        with ThreadPoolExecutor() as executor:
            executor.map(lambda p: process_group(*p), tasks)



def process_file():
    logging.info("Process file function called.")
    file_path = tk_interface.choose_file()
    if not file_path:
        logging.info("Файл не выбран")
        return

    try:
        main_wb = load_workbook(file_path)
        main_sheet = main_wb.active
    except Exception as e:
        logging.error(f"Ошибка при загрузке файла: {e}")
        messagebox.showerror("Ошибка", f"Не удалось загрузить файл: {e}")
        return
    tk_interface.progress_bar['value'] = 0

    output_folder = fm.create_output_folder(file_path=file_path)
    logging.info(f"Создана выходная папка: {output_folder}")

    genus_column_num, species_column_num = tk_interface.choose_columns()
    # if genus_column_num is None or species_column_num is None:
    #     return

    sorted_sheet = sort_group(file_path, main_sheet, genus_column_num, species_column_num)
    # with concurrent.futures.ThreadPoolExecutor() as executor:
    #     future = executor.submit(process_multi, sorted_sheet, genus_column_num, species_column_num, output_folder)
    #     future.add_done_callback(lambda f: tk_interface.update_progress(100))
    process_multi(sorted_sheet, genus_column_num, species_column_num, output_folder)
    tk_interface.update_progress(100)
    messagebox.showinfo("Завершение обработки", "Деление завершилось!")



if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    logging.info('Начинается вывод графического интерфейса')
    tk_interface = TKinter()
    tk_interface.set_process_callback(process_file)
    tk_interface.mainloop()