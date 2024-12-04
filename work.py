import time
from datetime import datetime
from tkinter import Tk, filedialog, Label, Button, ttk, simpledialog, Text
from tkinter.messagebox import askyesno
from openpyxl import load_workbook, Workbook, styles
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import os
from tqdm import tqdm
import csv
import random

data = 'Куратор;Начальник;Код;Ксюша;Антон;Загрузка на сайт'
file_number = 0
datemax = datetime.today().strftime('%Y%m%d%H%M')


def protection(sheet: str, col_letters: list, password: str):
    sheet.protection.sheet = True
    sheet.protection.password = password
    for col in col_letters:
        col_range = col.split(':')
        if len(col_range) == 2:
            start_col, end_col = col_range
        else:
            start_col = end_col = col_range[0]

            # Iterate over the columns in the range
            for column in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1):
                # Get the column letter
                col_letter = get_column_letter(column)
                # Protect the column
                for cell in sheet[col_letter]:
                    cell.protection = styles.Protection(locked=True)


def save_code_csv(data):
    # Split the data string into a list using ';' as the delimiter
    data_list = data.split(';')

    if os.path.exists(f'Рассылка{datemax}.csv'):
        mode = "a"
    else:
        mode = "w"

    # Open the CSV file in write mode
    with open(f'Рассылка{datemax}.csv', mode=mode, encoding='windows-1251', newline='') as file:
        # Create a CSV writer object
        writer = csv.writer(file, delimiter=';')

        # Write the data to the CSV file
        writer.writerow(data_list)


def create_output_folder(file_path):
    # Извлекаем имя файла из пути
    file_name = os.path.basename(file_path)
    # Извлекаем имя файла без расширения
    folder_name = os.path.splitext(file_name)[0]
    # Путь к папке для сохранения
    output_folder = os.path.join(os.path.expanduser("~"), "Downloads", folder_name)
    # Создаем папку
    os.makedirs(output_folder, exist_ok=True)
    return output_folder


def choose_file():
    root = Tk()
    root.withdraw()  # скрываем главное окно
    file_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Excel files", "*.xlsx")])
    return file_path


def update_progress(progress_value, max_value):
    percent = (progress_value / max_value) * 100 if max_value else 0
    progress_bar["value"] = percent
    status_label.config(text=f"Выполнено: {progress_value}/{max_value} ({percent:.2f}%)")
    root.update_idletasks()  # Обновляем интерфейс


def open_output_folder():
    """Открывает папку с результатами обработки."""
    if os.path.exists(output_folder):
        os.startfile(output_folder)
    else:
        status_label.config(text="Папка не существует.")


def close_window():
    """Закрывает окно приложения."""
    if askyesno("Закрыть окно", "Вы уверены, что хотите закрыть приложение?"):
        root.destroy()


def choose_columns():
    # Функция для преобразования буквы столбца в номер
    def column_letter_to_number(letter):
        return ord(letter.upper()) - 65

    curator_column_letter = simpledialog.askstring("Выбор столбца", "Введите букву столбца для куратора:")
    chief_column_letter = simpledialog.askstring("Выбор столбца", "Введите букву столбца для начальника:")

    # Преобразование буквы столбца в номер
    curator_column_number = column_letter_to_number(curator_column_letter)
    chief_column_number = column_letter_to_number(chief_column_letter)

    return curator_column_number, chief_column_number


def delete_colm(sheet, *colm_number):
    colm_index = [col_n - 1 for col_n in colm_number]
    for colm_index in sorted(colm_number, reverse=True):
        sheet.delete_cols(colm_index + 1)


def process_file():
    global output_folder, file_number
    time_right_now = datetime.today().strftime('%Y%m%d')
    save_code_csv(data)
    main_file_path = choose_file()
    if not main_file_path:
        status_label.config(text="Файл не выбран.")
        return

    output_folder = create_output_folder(main_file_path)

    main_wb = load_workbook(main_file_path)
    main_sheet = main_wb.active

    create_output_folder(main_file_path)

    dims = {}
    for row in main_sheet.rows:
        for cell in row:
            if cell.value:
                col_char = chr(64 + cell.column)
                if cell.column < 3:
                    min_wid = 7
                else:
                    min_wid = 17
                dims[col_char] = max((dims.get(cell.column, min_wid), len(str(cell.value))))

    curator_column_number, chief_column_number = choose_columns()

    # Создаем словарь для хранения всех файлов начальников
    chief_files = {}

    # Создаем все файлы начальников заранее
    for row in tqdm(range(2, main_sheet.max_row + 1), desc="Создание файлов начальников"):
        curator_value = main_sheet.cell(row=row, column=curator_column_number + 1).value
        chief_value = main_sheet.cell(row=row, column=chief_column_number + 1).value

        # Пропускаем пустые строки
        if not curator_value or not chief_value:
            continue

        # Санитизация имен куратора и начальника для создания папки и файла
        sanitized_curator = "".join(c for c in str(curator_value) if c.isalnum() or c in (' ', '_')).rstrip()
        sanitized_chief = "".join(c for c in str(chief_value) if c.isalnum() or c in (' ', '_')).rstrip()

        # Получаем путь к папке куратора
        curator_folder_path = os.path.join(output_folder, sanitized_curator)
        os.makedirs(curator_folder_path, exist_ok=True)

        # Получаем путь к файлу начальника
        chief_file_path = os.path.join(curator_folder_path, f'{sanitized_chief}.xlsx')

        # Создаем новый файл начальника, если его еще нет
        if chief_file_path not in chief_files:
            new_wb = Workbook()
            new_sheet = new_wb.active
            file_number = file_number + 1
            new_sheet.title = ((str(hex(int(time_right_now))[2:])) + str(file_number).zfill(3) + str(
                hex(random.randrange(1, 255))).zfill(2)).upper()
            save_code_csv(sanitized_curator + ';' + sanitized_chief + ';' + new_sheet.title)

            # Копируем стили заголовка из главного листа
            for col in range(1, main_sheet.max_column + 1):
                header_cell = main_sheet.cell(row=1, column=col)
                new_header_cell = new_sheet.cell(row=1, column=col, value=header_cell.value)

                if header_cell.has_style:
                    new_header_cell.font = copy(header_cell.font)
                    new_header_cell.border = copy(header_cell.border)
                    new_header_cell.fill = copy(header_cell.fill)
                    new_header_cell.number_format = copy(header_cell.number_format)
                    new_header_cell.protection = copy(header_cell.protection)
                    new_header_cell.alignment = copy(header_cell.alignment)

            # Сохраняем новый файл начальника в словаре
            chief_files[chief_file_path] = {'wb': new_wb, 'sheet': new_sheet}

    # Обработка строк и запись данных в соответствующие файлы начальников
    for row in tqdm(range(2, main_sheet.max_row + 1), desc="Обработка строк"):
        curator_value = main_sheet.cell(row=row, column=curator_column_number + 1).value
        chief_value = main_sheet.cell(row=row, column=chief_column_number + 1).value

        # Пропускаем пустые строки
        if not curator_value or not chief_value:
            continue

        # Получаем путь к папке куратора
        sanitized_curator = "".join(c for c in str(curator_value) if c.isalnum() or c in (' ', '_')).rstrip()
        curator_folder_path = os.path.join(output_folder, sanitized_curator)

        # Получаем путь к файлу начальника
        sanitized_chief = "".join(c for c in str(chief_value) if c.isalnum() or c in (' ', '_')).rstrip()
        chief_file_path = os.path.join(curator_folder_path, f'{sanitized_chief}.xlsx')

        # Получаем активный лист для текущего файла начальника
        current_wb = chief_files[chief_file_path]['wb']
        current_sheet = current_wb.active

        # Копируем данные и стили из главного листа в файл начальника
        new_row_index = current_sheet.max_row + 1

        for col in range(1, main_sheet.max_column + 1):
            cell = main_sheet.cell(row=row, column=col)
            new_cell = current_sheet.cell(row=new_row_index, column=col, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

        # # Устанавливаем ширину столбцов для текущего листа
        # for col_letter, max_width in max_widths.items():
        #     current_sheet.column_dimensions[col_letter].width = max_width

        for col, value in dims.items():
            current_sheet.column_dimensions[col].width = value

        delete_colm(current_sheet, curator_column_number, chief_column_number)

        if row % 20 == 0:
            update_progress(row, main_sheet.max_row)

        protection(current_sheet, )


    # Сохраняем все файлы начальников
    for chief_file_path, chief_file_info in chief_files.items():
        wb = chief_file_info['wb']
        wb.save(chief_file_path)

    # Устанавливаем прогресс в 100% и обновляем статус
    progress_bar["value"] = 100
    status_label.config(text="Завершено")

    # Закрываем главный файл
    main_wb.close()

    # После обработки файла активируем кнопку "Открыть папку"
    open_folder_button.config(state="normal")


root = Tk()
root.title("Процесс выполнения скрипта")
root.geometry("400x200")

status_label = Label(root, text="Нажмите кнопку 'Обработать файл' для начала обработки.")
status_label.pack(pady=10)

progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress_bar.pack(pady=10)

process_button = Button(root, text="Обработать файл", command=process_file)
process_button.pack(pady=5)

protection_area = Text(root, text="Столбцы для блокировки")
protection_area.pack(pady=5)

open_folder_button = Button(root, text="Открыть папку", command=open_output_folder, state="disabled")
open_folder_button.pack(pady=5)

close_button = Button(root, text="Закрыть", command=close_window)
close_button.pack(pady=5)

root.mainloop()
